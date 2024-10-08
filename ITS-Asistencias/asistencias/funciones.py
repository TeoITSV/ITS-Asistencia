from datetime import datetime, timedelta,time
import json
import locale
import openpyxl
from django.db.models import Min,Max
import pytz
from asistencias.models import *
from django.utils import timezone
import pandas as pd
from collections import defaultdict
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import calendar
from django.core.exceptions import ObjectDoesNotExist
from datetime import date
import re
def limpiar_nombre(nombre):
    # Convertir a minúsculas y eliminar espacios en blanco adicionales
    nombre = nombre.lower().strip()
    # Eliminar caracteres especiales y números
    nombre = re.sub(r'[^a-zA-Z\s]', '', nombre)
    return nombre
def calcularDiaNombre(dia):
    # recibe la fecha de la marca y obtiene que dia de la semana es, formateada para aceptar los dias que trabaja la bd, 
    # solo trabajamos de lunes  a viernes
    # dias_en_ingles = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    
    dias_en_espanol = Dia.objects.values_list('nombre', flat=True)
    traduccion_dias = {
    "Lunes": "Monday",
    "Martes": "Tuesday",
    "Miércoles": "Wednesday",
    "Jueves": "Thursday",
    "Viernes": "Friday",
    "Sábado": "Saturday",
    "Domingo": "Sunday",
    }
    dias_en_ingles = [traduccion_dias[dia] for dia in dias_en_espanol]

    diaNombre = dia.strftime("%A").capitalize()
    if diaNombre in dias_en_ingles:
        indice = dias_en_ingles.index(diaNombre)
        return dias_en_espanol[indice]
    elif diaNombre in dias_en_espanol:
        return diaNombre
    else:
        return None

def crearEmpleadoFila(row):
    empleado_nombre = limpiar_nombre(row['Nombre'])
    print(empleado_nombre)
    empleado_id = row['ID de usuario']
    print(f'Empleado: {empleado_nombre} - ID: {empleado_id}')

    if empleado_nombre and empleado_id and empleado_nombre != 'vacio' and empleado_id != 'vacio' or empleado_nombre=='':
        try:
            empleado, created = Empleado.objects.get_or_create(
                idEmpleado=empleado_id,
                defaults={'nombreCompleto': empleado_nombre}
            )
            print(empleado)
            return empleado
        except Exception as e:
            print(f'Error al crear o obtener el empleado: {e}')
            return None
    return None
def crearHorario(empleado):
    horariosViejos = Horario.objects.filter(empleado=empleado,esActual=True)
    #Borrado logico de los horarios viejos
    for horarioActual in horariosViejos:
        horarioActual.esActual = False
        horarioActual.fechaBajada = timezone.make_aware(datetime.now())
        horarioActual.save()
    #Creacion del horario nuevo
    horario = Horario(empleado=empleado)
    horario.save()
    return horario

def crearDiaHorario(horario,dia,valor_in,valor_out):
    if Horario.objects.filter(id=horario.id).exists():
        if valor_in != 'vacio' and valor_out != 'vacio':
            diaHorario = DiaHorario(horario=horario,diaNombre=dia,horaEntrada=valor_in,horaSalida=valor_out)
        else:
            diaHorario = DiaHorario(horario=horario,diaNombre=dia)
        diaHorario.save()
        return diaHorario
             
def crearHorarioFila(row,horarios_df, empleado):
    horario = crearHorario(empleado)
    for dia in Dia.objects.all():
        columna_in = dia.nombre
        valor_columna_in = row[dia.nombre]
        
        # Obtener la posición de la columna actual en el DataFrame
        posicion_columna_in = horarios_df.columns.get_loc(columna_in)
        
        # Obtener el nombre de la columna a la derecha
        columna_out = horarios_df.columns[posicion_columna_in + 1] if posicion_columna_in < len(horarios_df.columns) - 1 else None
        valor_columna_out = row[columna_out] if columna_out is not None else None
        #Aca se crea el Horario con sus diaHorario
        crearDiaHorario(horario,dia,valor_columna_in,valor_columna_out)


def crearMarca(diaHorario,empleado, fechaHora, tipoMarca):
    if tipoMarca:
        tipoMarca = TipoMarca.objects.filter(nombre=tipoMarca).first()
    else:
        tipoMarca = None
    marca, created = Marca.objects.get_or_create(diaHorario=diaHorario,empleado=empleado,fechaHora=timezone.make_aware(fechaHora),tipoMarca=tipoMarca)
    if created:
        marca.save()
        return marca, created
    return marca, created
def parseMarcas(df):
    
    # Crea un diccionario para almacenar las marcas de tiempo por fecha
    data_dict = defaultdict(lambda: {'Nombre': '', 'marcas': defaultdict(list)})

    # Itera sobre las filas del DataFrame
    for index, row in df.iterrows():
        id_empleado = row['ID de usuario']
        nombre_empleado = row['Nombre']
        fecha_hora = row['Fecha/Hora']

        # Convierte el objeto Timestamp a una cadena y luego extrae la fecha
        fecha = fecha_hora.strftime('%m/%d/%y').split()[0]

        # Almacena el nombre del empleado en el diccionario
        data_dict[id_empleado]['Nombre'] = nombre_empleado

        # Agrega la marca de tiempo a la lista correspondiente en el diccionario
        data_dict[id_empleado]['marcas'][fecha].append(str(fecha_hora))

    # Convierte el diccionario a una lista de diccionarios
    marcasEmpleadoPorDia = [{'ID de usuario': k, 'Nombre': v['Nombre'], 'marcas': v['marcas']} for k, v in data_dict.items()]
    return marcasEmpleadoPorDia
def crearMarcaEmpleado(empleado, dicMarcas):
    marcasCreadas = []
    try:
    # Se busca el horario actual del empleado
        print(f'Empleado: {empleado.nombreCompleto} - ID: {empleado.idEmpleado}')
        horario_actual = Horario.objects.filter(empleado=empleado,esActual=True).latest('fechaCreacion')
        print(f'Empleado: {empleado.nombreCompleto} - ID: {empleado.idEmpleado} - Horario: {horario_actual}')
        for dia, marcas in dicMarcas['marcas'].items():
            fecha_dt = datetime.strptime(dia, "%m/%d/%y")
            nombre_dia_semana = calcularDiaNombre(fecha_dt)
            if nombre_dia_semana is None:
                crearMarca(None,empleado,fecha_dt, None)
                continue
            # Se busca el día y horario correspondiente
            dia_horario_actual = DiaHorario.objects.filter(horario=horario_actual, diaNombre__nombre=nombre_dia_semana).first()
            print(f'Empleado: {empleado.nombreCompleto} - ID: {empleado.idEmpleado} - Dia: {dia}-{nombre_dia_semana} - Dia Horario: {dia_horario_actual}')
            # Convierte las marcas a objetos datetime
            marcas_dt = [datetime.strptime(marca, '%Y-%m-%d %H:%M:%S') for marca in marcas]
            # Obtén la hora de entrada y salida del horario actual
            #print(f'fecha_dt: {fecha_dt} - dia_horario_actual: {dia_horario_actual.horaEntrada}')
            if dia_horario_actual.horaEntrada is not None and dia_horario_actual.horaSalida is not None:
                hora_entrada_horario = datetime.combine(fecha_dt.date(), dia_horario_actual.horaEntrada)
                hora_salida_horario = datetime.combine(fecha_dt.date(), dia_horario_actual.horaSalida)
                marcaMax = max(marcas_dt)
                marcaMin = min(marcas_dt)
                tipoEntrada = None
                igual = False
                if marcaMax == marcaMin or len(marcas_dt) == 1:
                    igual = True
                    diferenciaHorarioIn = hora_entrada_horario - marcaMax
                    diferenciaHorarioOut = hora_salida_horario - marcaMax
                    if diferenciaHorarioIn < diferenciaHorarioOut:
                        tipoEntrada = 'In'
                    elif diferenciaHorarioIn > diferenciaHorarioOut:
                        tipoEntrada = 'Out'
                    else:
                        tipoEntrada = None
                    marca,created = crearMarca(dia_horario_actual,empleado, marcaMax, tipoEntrada)
                    if created == True:
                            marcasCreadas.append(marca)
                else:
                    marca1, created1 = crearMarca(dia_horario_actual,empleado, marcaMin, 'In')
                    marca2, created2 = crearMarca(dia_horario_actual,empleado, marcaMax, 'Out')
                    if created1 == True and created2 == True:
                        marcasCreadas.append(marca1)
                        marcasCreadas.append(marca2)
                for marca in marcas_dt:
                    if marca!= marcaMax and marca!= marcaMin and igual != True:
                        marca, created = crearMarca(dia_horario_actual,empleado, marca, None)
                        if created == True:
                            marcasCreadas.append(marca)
        return marcasCreadas
    except ObjectDoesNotExist:
        print(f'el empleado {empleado} no tiene horario cargado')
        return []
    except Exception as e:
        raise Exception(f'Error al crear las marcas de {empleado}: {e}')
    
def validarHorario(row, empleado, horarios_df):
    # implementar logica en un futuro para evitar entradas duplicadas el mismo dia
    horarios = Horario.objects.filter(empleado=empleado,esActual=True)
    if horarios.count() == 0:
        return True
    horario_dict = defaultdict(lambda: {'horario': None, 'diasHorario': []})

    # Llenar el diccionario con información sobre los horarios y sus días asociados
    for horario in horarios:
        diasHorario = DiaHorario.objects.filter(horario=horario)
        horario_dict[horario]['horario'] = horario
        horario_dict[horario]['diasHorario'].extend(diasHorario)
    # Imprimir los días de horario para cada horario
    for diaHorario in horario_dict[horario]['diasHorario']:
        dia = diaHorario.diaNombre.nombre
        if diaHorario.horaEntrada is None:
            diaHorario.horaEntrada = 'vacio'
        if diaHorario.horaSalida is None:
            diaHorario.horaSalida = 'vacio'
        columna_in = dia
        valor_columna_in = row[dia]
        if isinstance(valor_columna_in, datetime):
            valor_columna_in = valor_columna_in.strftime('%H:%M:%S')
        # Obtener la posición de la columna actual en el DataFrame
        posicion_columna_in = horarios_df.columns.get_loc(columna_in)
        
        # Obtener el nombre de la columna a la derecha
        columna_out = horarios_df.columns[posicion_columna_in + 1] if posicion_columna_in < len(horarios_df.columns) - 1 else None
        valor_columna_out = row[columna_out] if columna_out is not None else None
        if isinstance(valor_columna_out, datetime):
            valor_columna_out = valor_columna_out.strftime('%H:%M:%S')
        
        if str(valor_columna_in) != str(diaHorario.horaEntrada) or str(valor_columna_out) != str(diaHorario.horaSalida):
            return True
    return False
def leerPlanillaHorarios(horarioExcel):
    try:
        horarios_df = pd.read_excel(horarioExcel)
        horarios_df = horarios_df.where(pd.notna(horarios_df), "vacio")
        empleados = []
        for index, row in horarios_df.iloc[1:].iterrows():
            empleado = crearEmpleadoFila(row)
            if validarHorario(row, empleado, horarios_df):
                crearHorarioFila(row, horarios_df, empleado)
                empleados.append(empleado.nombreCompleto)
                yield True, f'Procesando horario para {empleado.nombreCompleto.title()}'
        if len(empleados) > 0:
            yield True, f"Se cargaron los horarios nuevos para los empleados: {', '.join(empleados)}"
        else:
            yield True, 'Se cargo la planilla exitosamente pero no hubo cambios con los horarios viejos'
    except FileNotFoundError:
        yield False, 'No se selecciono ningun archivo de horarios.'
    except Exception as e:
        yield False, f"Se produjo un error al leer el archivo Excel: {e}"
def leerPlanillaMarcas(marcasExcel):
    try:
        marcasCreadas = []
        df = pd.read_excel(marcasExcel)
        marcasEmpleadoPorDia = parseMarcas(df)
        for x in marcasEmpleadoPorDia:
            empleado = crearEmpleadoFila(x)
            marcasCreadas += crearMarcaEmpleado(empleado, x)
            yield True, f'Procesando marcas para {empleado.nombreCompleto.title()}'
        yield True, f'Se cargaron {len(marcasCreadas)} marcas'
    except FileNotFoundError:
        yield False, 'No se selecciono ningun archivo de marcas.'
    except Exception as e:
        yield False, f'Se produjo un error al leer el archivo Excel: {e}'
def calcInformeEmpleado(empleado,fechaInicio,fechaFin,margenEntrada):
    marcas = Marca.objects.filter(empleado=empleado,fechaHora__range=[fechaInicio, fechaFin]).order_by('fechaHora')
    llegadasTarde = []
    retirosFueradeHora = []
    zona_horaria_argentina = timezone.get_default_timezone()  # O usa timezone.pytz.timezone('America/Argentina/Buenos_Aires')

    for marca in marcas:
        marca.fechaHora = marca.fechaHora.replace(tzinfo=zona_horaria_argentina)
        if marca.tipoMarca is not None and marca.diaHorario is not None:
            if marca.tipoMarca.nombre == 'In':
                fecha_entrada = datetime.combine(marca.fechaHora.date(), marca.diaHorario.horaEntrada)
                
                # Convertir fecha_entrada a objeto consciente de la zona horaria
                fecha_entrada_con_zona = timezone.make_aware(fecha_entrada, zona_horaria_argentina)

                # Sumar el margen directamente a la fecha y hora de entrada
                fechaHoraInHorario = fecha_entrada_con_zona + timedelta(minutes=margenEntrada.minute, seconds=margenEntrada.second)
                
                if (marca.fechaHora - timedelta(minutes=1)).replace(second=0) > fechaHoraInHorario:
                    diferencia_tiempo = marca.fechaHora - fechaHoraInHorario
                    #print(f'Empleado: {empleado.nombreCompleto} - ID: {empleado.idEmpleado}- marca: {marca.fechaHora} - horaIn: {fechaHoraInHorario} - Diferencia: {diferencia_tiempo}')
                    marca.diferenciaFalta = diferencia_tiempo
                    llegadasTarde.append(marca)
                    #print(f'Empleado: {empleado.nombreCompleto} - ID: {empleado.idEmpleado} ')
            elif marca.tipoMarca.nombre == 'Out':
                fecha_salida = datetime.combine(marca.fechaHora.date(), marca.diaHorario.horaSalida)
                
                # Convertir fecha_entrada a objeto consciente de la zona horaria
                fecha_salida_con_zona = timezone.make_aware(fecha_salida, zona_horaria_argentina)

                # Sumar el margen directamente a la fecha y hora de entrada
                fechaHoraOutHorario = fecha_salida_con_zona + timedelta(minutes=margenEntrada.minute, seconds=margenEntrada.second)
                if (marca.fechaHora + timedelta(minutes= 1)).replace(second=0) < fechaHoraOutHorario:
                    diferencia_tiempo = marca.fechaHora - fechaHoraOutHorario
                    marca.diferenciaFal = diferencia_tiempo
                    retirosFueradeHora.append(marca)
    return llegadasTarde , retirosFueradeHora
def calcStats(date_min,date_max):
    if date_min is None or date_max is None:
        date_max = Marca.objects.aggregate(max_fecha=Max('fechaHora'))['max_fecha']
        date_min = Marca.objects.aggregate(min_fecha=Min('fechaHora'))['min_fecha']
    margenEntrada = time(0,1,0)
    total_retrasos = 0
    total_salidas_tempranas = 0
    
    for empleado in Empleado.objects.all():
        retrasos_empleado,empleado_salidas_tempranas = calcInformeEmpleado(empleado,date_min,date_max,margenEntrada)
        total_retrasos+=len(retrasos_empleado)
        total_salidas_tempranas+=len(empleado_salidas_tempranas)
    return total_retrasos, total_salidas_tempranas

def ajustar_ancho_columnas(hoja):
    for col in hoja.columns:
        max_length = 0
        column = col[0].column_letter  # Obtiene la letra de la columna (A, B, C, etc.)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        hoja.column_dimensions[column].width = adjusted_width

def descargarInformeExcel(empleados):
    # Crear un nuevo libro de trabajo
    workbook = Workbook()
    archivo = f'attachment; filename=registro{datetime.now().strftime("%Y-%m-%d_%Hhs%Mm%Ss")}.xlsx'
    # Seleccionar la hoja activa
    sheet = workbook.active

    # Escribir datos en algunas celdas
    columnas = ['ID Empleado', 'Empleado', 'Fecha','In/Out','Horario','Marca','Diferencia']
    
    sheet.append(columnas)
    for data_row in empleados:
        for llegadaTarde in data_row.llegadasTarde:
            fila = [data_row.idEmpleado, data_row.nombreCompleto, llegadaTarde.fechaHora.strftime('%d/%m/%Y'),llegadaTarde.tipoMarca.nombre,llegadaTarde.diaHorario.horaEntrada.strftime('%H:%M:%S'),llegadaTarde.fechaHora.strftime('%H:%M:%S'),llegadaTarde.diferenciaFalta]
            sheet.append(fila)
        for retirosFueradeHora in data_row.retirosFueradeHora:
            fila = [data_row.idEmpleado, data_row.nombreCompleto, retirosFueradeHora.fechaHora.strftime('%d/%m/%Y'),retirosFueradeHora.tipoMarca.nombre,retirosFueradeHora.diaHorario.horaSalida.strftime('%H:%M:%S'),retirosFueradeHora.fechaHora.strftime('%H:%M:%S'),retirosFueradeHora.diferenciaFal]
            sheet.append(fila)

     # Ajustar el formato de las celdas para que se vean mejor
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')  # Alineación centrada

    # Aplicar autofiltros
    sheet.auto_filter.ref = sheet.dimensions
    ajustar_ancho_columnas(sheet)
    # Crear una respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = archivo

    # Guardar el libro de trabajo en el flujo de respuesta
    workbook.save(response)

    return response
def generarInforme(form):
    fechaInicio = form.cleaned_data['fechaInicio']
    fechaFin = form.cleaned_data['fechaFin']
    minutos = form.cleaned_data['minutos'] or 0
    segundos = form.cleaned_data['segundos'] or 0

        # Crea un objeto timedelta con los valores proporcionados
    margenEntrada = time(0,minutos,segundos)
    selectAll = form.cleaned_data['selectAll']
    if selectAll == 'all':
        empleados = Empleado.objects.all()
    else:
        empleados = [Empleado.objects.get(idEmpleado=empleado) for empleado in form.cleaned_data['empleados']]
    for empleado in empleados:
        llegadasTarde, retirosFueradeHora = calcInformeEmpleado(empleado,fechaInicio,fechaFin,margenEntrada)
        empleado.llegadasTarde = llegadasTarde
        empleado.retirosFueradeHora = retirosFueradeHora
    return descargarInformeExcel(empleados)
class PieChartHome():
    def calcStats(self):
        min_fecha = Marca.objects.aggregate(min_fecha=Min('fechaHora'))['min_fecha']
        max_fecha = Marca.objects.aggregate(max_fecha=Max('fechaHora'))['max_fecha']
        if min_fecha is not None:
            minAnio = min_fecha.year
        else:
            return []

        if max_fecha is not None:
            maxAnio = max_fecha.year
        else:
            return []
        anios = [anio for anio in range(minAnio, maxAnio + 1)]
        stats = []
        labels = ["En horario", "Anticipados", "Retrasos"]
        for anio in anios:
            cantidad = self.calcAnio(anio)
            stats.append({
                'year': anio,
                'total': sum(cantidad),
                'data':{
                    'labels': labels,
                    'data': cantidad
                }
            })
        return stats
    def calcAnio(self,year):
        date_min = date(year, 1, 1)
        date_max = date(year, 12, 31)
        margenEntrada = time(0,1,0)
        total_retrasos = 0
        total_salidas_tempranas = 0
        total_marcas = Marca.objects.filter(fechaHora__range=[date_min, date_max]).count()
        for empleado in Empleado.objects.all():
            retrasos_empleado,empleado_salidas_tempranas = calcInformeEmpleado(empleado,date_min,date_max,margenEntrada)
            total_retrasos+=len(retrasos_empleado)
            total_salidas_tempranas+=len(empleado_salidas_tempranas)
        totalEnHorario = total_marcas - total_retrasos - total_salidas_tempranas
        return [totalEnHorario, total_salidas_tempranas, total_retrasos]
class HistogramaHome():
    def marcasxmes(self,year):
        labels = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        data_list = [Marca.objects.filter(fechaHora__month=mes).filter(fechaHora__year=year).count() for mes in range(1, 13)]
        return labels, data_list
    def marcasAnio(self):
        min_fecha = Marca.objects.aggregate(min_fecha=Min('fechaHora'))['min_fecha']
        max_fecha = Marca.objects.aggregate(max_fecha=Max('fechaHora'))['max_fecha']
        if min_fecha is not None:
            minAnio = min_fecha.year
        else:
            minAnio = 2023

        if max_fecha is not None:
            maxAnio = max_fecha.year
        else:
            maxAnio = date.today().year
        anios = [anio for anio in range(minAnio, maxAnio + 1)]
        tablaFrecuencias = []
        for anio in anios:
            mes, frecuencia = self.marcasxmes(anio)
            tablaFrecuencias.append({
                'year': anio,
                'meses':{
                    'labels': mes,
                    'data': frecuencia
                }
            })
        return tablaFrecuencias
    def faltasAnio(self):
        min_fecha = Marca.objects.aggregate(min_fecha=Min('fechaHora'))['min_fecha']
        max_fecha = Marca.objects.aggregate(max_fecha=Max('fechaHora'))['max_fecha']
        if min_fecha is not None:
            minAnio = min_fecha.year
        else:
            minAnio = 2023

        if max_fecha is not None:
            maxAnio = max_fecha.year
        else:
            maxAnio = date.today().year
        anios = [anio for anio in range(minAnio, maxAnio + 1)]
        tablaFrecuencias = []
        for anio in anios:
            mes, frecuenciaRetrasos, frecuenciaRetiros = self.faltasxmes(anio)
            tablaFrecuencias.append({
                'year': anio,
                'meses':{
                    'labels': mes,
                    'retrasos': frecuenciaRetrasos,
                    'retirosAnticipadas': frecuenciaRetiros
                }
            })
        return tablaFrecuencias
    def faltasxmes(self,year):
        labels = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        retrasosList = []
        retirosList = []
        for mes in range(1, 13):
            primer_dia = date(year, mes, 1)
            ultimo_dia = date(year, mes, calendar.monthrange(year, mes)[1])
            retrasos, retiros = calcStats(primer_dia,ultimo_dia)
            retrasosList.append(retrasos)
            retirosList.append(retiros)

        return labels, retrasosList, retirosList


